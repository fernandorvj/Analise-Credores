"""Exportação para Excel dos resultados da Calculadora — Simulador de
Financiamento, Simulação Balão, Calculadora de VPL e Comparação de Cenários.

Segue o mesmo padrão do exportador de Credores (``pandas.ExcelWriter`` +
pós-formatação via ``openpyxl``), mas isolado num módulo próprio — não
importa nem altera `src/exportar_excel.py`, para manter a Calculadora
totalmente desacoplada do módulo Credores.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.calculadora.models import Cenario, FluxoItem, ResultadoFinanciamento, ResultadoVPL
from src.utils import formatar_moeda, formatar_percentual

_FORMATO_MOEDA = "R$ #,##0.00"
_FORMATO_PERCENTUAL = "0.00%"

COLUNAS_MOEDA_CALC = {
    "Saldo Inicial",
    "Juros",
    "Amortização",
    "Valor Parcela",
    "Saldo Final",
    "Valor",
    "Saldo Devedor",
    "Valor Presente",
}
COLUNAS_PERCENTUAL_CALC: set[str] = set()


def _aplicar_formatos(ws: Worksheet, df: pd.DataFrame) -> None:
    """Aplica formato de moeda/percentual (por nome de coluna) e ajusta a
    largura das colunas de uma aba — mesma convenção do exportador de
    Credores, reimplementada aqui para manter a Calculadora isolada.
    """
    for indice, coluna in enumerate(df.columns, start=1):
        letra = get_column_letter(indice)
        if coluna in COLUNAS_MOEDA_CALC:
            formato = _FORMATO_MOEDA
        elif coluna in COLUNAS_PERCENTUAL_CALC:
            formato = _FORMATO_PERCENTUAL
        else:
            formato = None
        if formato:
            for celula in ws[letra][1:]:
                celula.number_format = formato
        maior_valor = max([len(str(coluna))] + [len(str(v)) for v in df[coluna].astype(str)])
        ws.column_dimensions[letra].width = min(max(maior_valor + 2, 10), 45)
    for celula in ws[1]:
        celula.font = Font(bold=True)


def _linha_resumo(indicador: str, valor: str) -> dict:
    return {"Indicador": indicador, "Valor": valor}


def _df_cronograma_financiamento(resultado: ResultadoFinanciamento) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Nº": p.numero,
                "Data": p.data,
                "Carência": "Sim" if p.carencia else "Não",
                "Saldo Inicial": float(p.saldo_inicial),
                "Juros": float(p.juros),
                "Amortização": float(p.amortizacao),
                "Valor Parcela": float(p.valor_parcela),
                "Saldo Final": float(p.saldo_final),
            }
            for p in resultado.parcelas
        ]
    )


def exportar_excel_financiamento(resultado: ResultadoFinanciamento, caminho_saida: str | Path) -> Path:
    """Exporta o cronograma do Simulador de Financiamento (abas Resumo e
    Cronograma) para um arquivo .xlsx."""
    caminho_saida = Path(caminho_saida)
    p = resultado.parametros

    df_resumo = pd.DataFrame(
        [
            _linha_resumo("Sistema de Amortização", p.sistema.value),
            _linha_resumo("Regime de Juros", p.regime.value),
            _linha_resumo("Valor Financiado", formatar_moeda(float(p.valor_financiado))),
            _linha_resumo("Valor de Entrada", formatar_moeda(float(p.valor_entrada))),
            _linha_resumo("Taxa Informada", formatar_percentual(float(p.taxa))),
            _linha_resumo("Periodicidade da Taxa", p.periodicidade_taxa.value),
            _linha_resumo("Periodicidade das Parcelas", p.periodicidade_parcela.value),
            _linha_resumo("Prazo (parcelas)", str(p.prazo)),
            _linha_resumo("Carência (parcelas)", str(p.carencia)),
            _linha_resumo("Taxa Periódica Efetiva", formatar_percentual(float(resultado.taxa_periodica))),
            _linha_resumo("Juros Totais", formatar_moeda(float(resultado.juros_totais))),
            _linha_resumo("Total Pago", formatar_moeda(float(resultado.total_pago))),
        ]
    )
    df_cronograma = _df_cronograma_financiamento(resultado)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_cronograma.to_excel(writer, sheet_name="Cronograma", index=False)
        _aplicar_formatos(writer.sheets["Cronograma"], df_cronograma)
        for celula in writer.sheets["Resumo"][1]:
            celula.font = Font(bold=True)

    return caminho_saida


def _df_fluxo(fluxo: list[FluxoItem]) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "Data": item.data,
                "Descrição": item.descricao,
                "Tipo": item.tipo.value,
                "Valor": float(item.valor),
                "Juros": float(item.juros) if item.juros is not None else None,
                "Amortização": float(item.amortizacao) if item.amortizacao is not None else None,
                "Saldo Devedor": float(item.saldo_devedor) if item.saldo_devedor is not None else None,
                "Valor Presente": float(item.valor_presente) if item.valor_presente is not None else None,
            }
            for item in fluxo
        ]
    )


def exportar_excel_fluxo(fluxo: list[FluxoItem], caminho_saida: str | Path, titulo_aba: str = "Fluxo") -> Path:
    """Exporta um fluxo de caixa livre (Simulação Balão / Editor de Fluxo) para .xlsx."""
    caminho_saida = Path(caminho_saida)
    df = _df_fluxo(fluxo)
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=titulo_aba[:31], index=False)
        _aplicar_formatos(writer.sheets[titulo_aba[:31]], df)
    return caminho_saida


def exportar_excel_vpl(resultado: ResultadoVPL, caminho_saida: str | Path) -> Path:
    """Exporta o resultado da Calculadora de VPL (abas Resumo e Fluxo
    Descontado) para um arquivo .xlsx."""
    caminho_saida = Path(caminho_saida)
    p = resultado.parametros

    linhas_resumo = [
        _linha_resumo("Valor do Crédito", formatar_moeda(float(p.valor_credito))),
        _linha_resumo("Deságio do Plano de RJ", formatar_percentual(float(p.desagio))),
        _linha_resumo("Valor a Receber pelo Plano", formatar_moeda(float(p.valor_credito * (1 - p.desagio)))),
        _linha_resumo("Valor de Compra", formatar_moeda(float(p.valor_compra))),
        _linha_resumo("Taxa de Desconto (a.a.)", formatar_percentual(float(p.taxa_desconto_anual))),
        _linha_resumo("Origem da Taxa de Desconto", p.origem_taxa_desconto),
        _linha_resumo("Correção Monetária (a.a.)", formatar_percentual(float(p.correcao_monetaria_anual))),
        _linha_resumo("Valor Futuro", formatar_moeda(float(resultado.valor_futuro))),
        _linha_resumo("VPL (valor presente do fluxo)", formatar_moeda(float(resultado.vpl))),
        _linha_resumo("Ganho Líquido (VPL − Compra)", formatar_moeda(float(resultado.ganho_liquido))),
        _linha_resumo(
            "TIR (a.a.)", formatar_percentual(float(resultado.tir_anual)) if resultado.tir_anual is not None else "-"
        ),
        _linha_resumo(
            "Taxa Efetiva (a.a.)",
            formatar_percentual(float(resultado.taxa_efetiva_anual)) if resultado.taxa_efetiva_anual is not None else "-",
        ),
        _linha_resumo("Payback", str(resultado.payback_data) if resultado.payback_data else "Não atingido"),
        _linha_resumo("ROI", formatar_percentual(float(resultado.roi)) if resultado.roi is not None else "-"),
        _linha_resumo(
            "Rentabilidade", formatar_percentual(float(resultado.rentabilidade)) if resultado.rentabilidade is not None else "-"
        ),
        _linha_resumo("Margem", formatar_percentual(float(resultado.margem)) if resultado.margem is not None else "-"),
        _linha_resumo("Spread (a.a.)", formatar_percentual(float(resultado.spread)) if resultado.spread is not None else "-"),
    ]
    df_resumo = pd.DataFrame(linhas_resumo)
    df_fluxo = _df_fluxo(resultado.fluxo_descontado)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)
        df_fluxo.to_excel(writer, sheet_name="Fluxo Descontado", index=False)
        _aplicar_formatos(writer.sheets["Fluxo Descontado"], df_fluxo)
        for celula in writer.sheets["Resumo"][1]:
            celula.font = Font(bold=True)

    return caminho_saida


def exportar_excel_comparacao(cenarios: list[Cenario], caminho_saida: str | Path) -> Path:
    """Exporta uma tabela comparativa de múltiplos cenários (financiamento
    e/ou VPL) para um arquivo .xlsx com uma aba única."""
    caminho_saida = Path(caminho_saida)
    linhas = []
    for cenario in cenarios:
        if cenario.tipo == "vpl":
            r: ResultadoVPL = cenario.resultado  # type: ignore[assignment]
            linhas.append(
                {
                    "Cenário": cenario.nome,
                    "Tipo": "VPL",
                    "VPL": float(r.vpl),
                    "TIR (a.a.)": float(r.tir_anual) if r.tir_anual is not None else None,
                    "ROI": float(r.roi) if r.roi is not None else None,
                    "Payback": str(r.payback_data) if r.payback_data else "-",
                    "Rentabilidade": float(r.rentabilidade) if r.rentabilidade is not None else None,
                }
            )
        else:
            f: ResultadoFinanciamento = cenario.resultado  # type: ignore[assignment]
            linhas.append(
                {
                    "Cenário": cenario.nome,
                    "Tipo": "Financiamento",
                    "Valor Parcela": float(f.valor_parcela_regular) if f.valor_parcela_regular is not None else None,
                    "Juros Totais": float(f.juros_totais),
                    "Total Pago": float(f.total_pago),
                }
            )
    df = pd.DataFrame(linhas)
    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Comparação", index=False)
        for celula in writer.sheets["Comparação"][1]:
            celula.font = Font(bold=True)
    return caminho_saida
