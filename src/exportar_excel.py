"""Exportação da análise de credores para uma planilha Excel (.xlsx) com múltiplas abas."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import configurar_logging
from src import analise_quorum, estrategia
from src.models import ResultadoExtracao
from src.utils import COLUNAS_MOEDA_PADRAO, COLUNAS_PERCENTUAL_PADRAO

logger = configurar_logging()

_FORMATO_MOEDA = 'R$ #,##0.00'
_FORMATO_PERCENTUAL = '0.00%'


def _aplicar_formatos(ws: Worksheet, df: pd.DataFrame) -> None:
    """Aplica formato de moeda/percentual e ajusta largura das colunas de uma aba."""
    for indice, coluna in enumerate(df.columns, start=1):
        letra = get_column_letter(indice)
        if coluna in COLUNAS_MOEDA_PADRAO:
            formato = _FORMATO_MOEDA
        elif coluna in COLUNAS_PERCENTUAL_PADRAO:
            formato = _FORMATO_PERCENTUAL
        else:
            formato = None

        if formato:
            for celula in ws[letra][1:]:
                celula.number_format = formato

        maior_valor = max([len(str(coluna))] + [len(str(v)) for v in df[coluna].astype(str)])
        ws.column_dimensions[letra].width = min(max(maior_valor + 2, 10), 45)

    for celula in ws[1]:
        celula.font = Font(bold=True)


def exportar_excel(
    resultado: ResultadoExtracao,
    caminho_saida: str | Path,
    top_n_ranking: int = 50,
) -> Path:
    """Gera um .xlsx com abas: Tabela Analítica, Resumo por Classe, Ranking,
    Credores Estratégicos, Simulações de Quórum e Pendências de Revisão.
    """
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    df_analitica = analise_quorum.tabela_analitica(resultado.credores)
    df_resumo_classe = analise_quorum.resumo_por_classe(resultado.credores)
    df_ranking = analise_quorum.ranking_maiores_credores(resultado.credores, top_n=top_n_ranking)
    df_estrategicos = estrategia.credores_estrategicos(resultado.credores)

    simulacoes = estrategia.simular_formacao_quorum(resultado.credores)
    df_simulacoes = estrategia.tabela_simulacoes(simulacoes)

    pendencias = resultado.credores_para_revisar + resultado.credores_com_erro
    df_pendencias = pd.DataFrame([c.to_dict() for c in pendencias])

    abas: dict[str, pd.DataFrame] = {
        "Tabela Analitica": df_analitica,
        "Resumo por Classe": df_resumo_classe,
        "Ranking Maiores Credores": df_ranking,
        "Credores Estrategicos": df_estrategicos,
        "Simulacoes de Quorum": df_simulacoes,
        "Pendencias de Revisao": df_pendencias,
    }

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        for nome_aba, df in abas.items():
            (df if not df.empty else pd.DataFrame({"Aviso": ["Sem dados"]})).to_excel(
                writer, sheet_name=nome_aba, index=False
            )
            if not df.empty:
                _aplicar_formatos(writer.sheets[nome_aba], df)

    logger.info("Excel exportado em '%s' (%d credores).", caminho_saida, resultado.total_credores)
    return caminho_saida
